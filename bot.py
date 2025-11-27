import os
from datetime import datetime

from dotenv import load_dotenv
import telebot
from telebot import types
from openpyxl import Workbook, load_workbook

# ================= НАСТРОЙКИ =====================

# Загружаем .env и токен
load_dotenv()
TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")

print("Loaded TELEGRAM_BOT_TOKEN:", TOKEN)

if not TOKEN:
    raise RuntimeError("Не найден TELEGRAM_BOT_TOKEN в файле .env")

bot = telebot.TeleBot(TOKEN)

# Имя Excel-файла очереди заказов
EXCEL_FILE = "orders.xlsx"
MAX_ITEMS_PER_ORDER = 10  # сколько позиций товара максимум пишем в строку

# Каталог товаров (пример — под себя можешь поменять)
# ВАЖНО: пути к картинкам существуют в папке images/
PRODUCTS = [
    {
        "id": 1,
        "name": "Фигурка дракона",
        "price": 500,
        "model": "dragon.stl",
        "description": "Дракон 10 см, PLA-пластик.",
        "image": "images/dragon.jpg",
    },
    {
        "id": 2,
        "name": "Держатель для телефона",
        "price": 300,
        "model": "phone_holder.stl",
        "description": "Универсальный держатель для смартфона.",
        "image": "images/phone_holder.jpg",
    },
    {
        "id": 3,
        "name": "Ключница настенная",
        "price": 450,
        "model": "key_holder.stl",
        "description": "Настенная ключница на 5 крючков.",
        "image": "images/key_holder.jpg",
    },
]

# ================= СОСТОЯНИЯ ПОЛЬЗОВАТЕЛЕЙ =================

user_carts = {}       # user_id -> [ {name, qty, price, model}, ... ]
user_states = {}      # user_id -> state (None, waiting_qty, waiting_fio, waiting_phone)
pending_product = {}  # user_id -> product_id
checkout_data = {}    # user_id -> {"fio": ..., "phone": ...}

# Очередь заказов для оператора
orders_queue = []     # список словарей с заказами


# ================= РАБОТА С EXCEL =========================

def init_workbook():
    """Создаем Excel с заголовками, если его еще нет."""
    if os.path.exists(EXCEL_FILE):
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Заказы"

    headers = ["Дата заказа", "ФИО", "Телефон"]

    # Далее блоки по 3 столбца для каждой позиции заказа
    for i in range(1, MAX_ITEMS_PER_ORDER + 1):
        headers.extend([
            f"Имя товара {i}",
            f"Кол-во товара {i} (шт)",
            f"Модель {i} (имя_товара.stl)",
        ])

    ws.append(headers)
    wb.save(EXCEL_FILE)


def save_order_to_excel(fio: str, phone: str, items: list):
    """
    Сохранение заказа в Excel-файл.
    items: список dict с ключами name, qty, model
    """
    init_workbook()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    date_str = datetime.now().strftime("%d.%m.%Y %H:%M")
    row = [date_str, fio, phone]

    # Гарантируем фиксированное количество колонок на строку
    for i in range(MAX_ITEMS_PER_ORDER):
        if i < len(items):
            item = items[i]
            row.extend([item["name"], item["qty"], item["model"]])
        else:
            row.extend(["", "", ""])

    ws.append(row)
    wb.save(EXCEL_FILE)


# ================= ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ =================

def get_product_by_id(prod_id: int):
    for p in PRODUCTS:
        if p["id"] == prod_id:
            return p
    return None


def get_cart(user_id: int):
    return user_carts.get(user_id, [])


def add_to_cart(user_id: int, product: dict, qty: int):
    cart = user_carts.setdefault(user_id, [])
    cart.append({
        "name": product["name"],
        "qty": qty,
        "price": product["price"],
        "model": product["model"],
    })


def format_cart_text(user_id: int) -> str:
    cart = get_cart(user_id)
    if not cart:
        return "🛒 Ваша корзина пуста."

    total = 0
    lines = []
    for i, item in enumerate(cart, start=1):
        line_sum = item["price"] * item["qty"]
        total += line_sum
        lines.append(
            f"{i}. {item['name']} — {item['qty']} шт × {item['price']} ₽ = {line_sum} ₽"
        )

    lines.append(f"\nИтого: {total} ₽")
    return "\n".join(lines)


def main_menu_keyboard():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
    kb.add(types.KeyboardButton("Каталог товаров"),
           types.KeyboardButton("Корзина"))
    return kb


def cart_keyboard():
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
    kb.add(types.KeyboardButton("Оформить заказ"),
           types.KeyboardButton("Очистить корзину"))
    kb.add(types.KeyboardButton("Каталог товаров"))
    return kb


def send_catalog_cards(chat_id: int):
    """Показать карточки товаров: фото + описание + кнопка."""
    for p in PRODUCTS:
        kb = types.InlineKeyboardMarkup()
        kb.add(types.InlineKeyboardButton(
            text=f"Добавить в корзину",
            callback_data=f"add_{p['id']}"
        ))

        caption = (
            f"<b>{p['name']}</b>\n"
            f"Цена: {p['price']} ₽\n"
            f"{p['description']}\n"
            f"Модель: <code>{p['model']}</code>"
        )

        image_path = p.get("image")
        if image_path and os.path.exists(image_path):
            with open(image_path, "rb") as img:
                bot.send_photo(
                    chat_id,
                    img,
                    caption=caption,
                    parse_mode="HTML",
                    reply_markup=kb
                )
        else:
            # Если картинка не найдена — просто текст
            bot.send_message(
                chat_id,
                caption,
                parse_mode="HTML",
                reply_markup=kb
            )


# ================== ХЕНДЛЕРЫ ПОЛЬЗОВАТЕЛЕЙ ==================

@bot.message_handler(commands=["start"])
def handle_start(message: types.Message):
    user_id = message.from_user.id
    user_states[user_id] = None
    user_carts[user_id] = []

    bot.send_message(
        message.chat.id,
        "Привет! 👋\n"
        "Я бот для оформления заказов на 3D-печать.\n\n"
        "Я умею:\n"
        "• показывать каталог товаров (карточки с фото);\n"
        "• добавлять товары в корзину;\n"
        "• оформлять заказ (ФИО + телефон);\n"
        "• записывать заказ в Excel и в очередь.\n\n"
        "Нажми «Каталог товаров», чтобы посмотреть продукцию.",
        reply_markup=main_menu_keyboard(),
    )


@bot.message_handler(func=lambda m: m.text == "Каталог товаров")
def handle_catalog(message: types.Message):
    send_catalog_cards(message.chat.id)


@bot.callback_query_handler(func=lambda c: c.data.startswith("add_"))
def handle_add_product(call: types.CallbackQuery):
    user_id = call.from_user.id
    prod_id = int(call.data.split("_")[1])
    product = get_product_by_id(prod_id)

    if not product:
        bot.answer_callback_query(call.id, "Товар не найден.")
        return

    pending_product[user_id] = prod_id
    user_states[user_id] = "waiting_qty"

    bot.answer_callback_query(call.id)
    bot.send_message(
        call.message.chat.id,
        f"Сколько штук товара «{product['name']}» добавить в корзину? Введите число."
    )


@bot.message_handler(func=lambda m: user_states.get(m.from_user.id) == "waiting_qty")
def handle_quantity(message: types.Message):
    user_id = message.from_user.id
    text = message.text.strip()

    if not text.isdigit() or int(text) <= 0:
        bot.send_message(
            message.chat.id,
            "Пожалуйста, введите целое положительное число.",
        )
        return

    qty = int(text)
    prod_id = pending_product.get(user_id)
    product = get_product_by_id(prod_id)

    if not product:
        bot.send_message(
            message.chat.id,
            "Ошибка: товар не найден. Попробуйте снова через каталог.",
            reply_markup=main_menu_keyboard(),
        )
        user_states[user_id] = None
        pending_product.pop(user_id, None)
        return

    add_to_cart(user_id, product, qty)
    user_states[user_id] = None
    pending_product.pop(user_id, None)

    bot.send_message(
        message.chat.id,
        f"✅ Добавлено в корзину: {product['name']} — {qty} шт.\n\n"
        f"{format_cart_text(user_id)}",
        reply_markup=cart_keyboard(),
    )


@bot.message_handler(func=lambda m: m.text == "Корзина")
def handle_cart(message: types.Message):
    user_id = message.from_user.id
    text = format_cart_text(user_id)
    reply_kb = cart_keyboard() if get_cart(user_id) else main_menu_keyboard()
    bot.send_message(message.chat.id, text, reply_markup=reply_kb)


@bot.message_handler(func=lambda m: m.text == "Очистить корзину")
def handle_clear_cart(message: types.Message):
    user_id = message.from_user.id
    user_carts[user_id] = []
    bot.send_message(
        message.chat.id,
        "Корзина очищена.",
        reply_markup=main_menu_keyboard(),
    )


@bot.message_handler(func=lambda m: m.text == "Оформить заказ")
def handle_checkout_start(message: types.Message):
    user_id = message.from_user.id
    cart = get_cart(user_id)

    if not cart:
        bot.send_message(
            message.chat.id,
            "Ваша корзина пуста. Сначала добавьте товары из каталога.",
            reply_markup=main_menu_keyboard(),
        )
        return

    user_states[user_id] = "waiting_fio"
    checkout_data[user_id] = {}
    bot.send_message(
        message.chat.id,
        "Для оформления заказа введите, пожалуйста, ваше ФИО полностью:"
    )


@bot.message_handler(func=lambda m: user_states.get(m.from_user.id) == "waiting_fio")
def handle_checkout_fio(message: types.Message):
    user_id = message.from_user.id
    fio = message.text.strip()

    if len(fio.split()) < 2:
        bot.send_message(
            message.chat.id,
            "Пожалуйста, введите фамилию и имя (можно с отчеством).",
        )
        return

    checkout_data[user_id]["fio"] = fio
    user_states[user_id] = "waiting_phone"
    bot.send_message(
        message.chat.id,
        "Введите, пожалуйста, ваш номер телефона:"
    )


@bot.message_handler(func=lambda m: user_states.get(m.from_user.id) == "waiting_phone")
def handle_checkout_phone(message: types.Message):
    user_id = message.from_user.id
    phone = message.text.strip()

    if len(phone) < 6:
        bot.send_message(
            message.chat.id,
            "Номер телефона выглядит слишком коротким. Попробуйте еще раз:",
        )
        return

    fio = checkout_data[user_id]["fio"]
    checkout_data[user_id]["phone"] = phone
    cart = get_cart(user_id)

    # 1) сохраняем в Excel
    save_order_to_excel(fio=fio, phone=phone, items=cart)

    # 2) добавляем в очередь заказов
    order_entry = {
        "timestamp": datetime.now().strftime("%d.%m.%Y %H:%M"),
        "fio": fio,
        "phone": phone,
        "items": cart.copy(),
    }
    orders_queue.append(order_entry)

    # 3) очищаем данные пользователя
    user_states[user_id] = None
    user_carts[user_id] = []
    checkout_data.pop(user_id, None)

    bot.send_message(
        message.chat.id,
        "🎉 Спасибо! Ваш заказ оформлен.\n"
        "Он добавлен в очередь и записан в Excel.",
        reply_markup=main_menu_keyboard(),
    )


# ================== ОЧЕРЕДЬ ЗАКАЗОВ (для оператора) ==================

@bot.message_handler(commands=["queue"])
def handle_queue(message: types.Message):
    if not orders_queue:
        bot.send_message(message.chat.id, "Очередь заказов пуста.")
        return

    text = "📦 <b>Очередь заказов:</b>\n\n"
    for i, o in enumerate(orders_queue, start=1):
        text += (
            f"{i}. <b>{o['fio']}</b> ({o['phone']}) — {o['timestamp']}\n"
            f"Товаров: {len(o['items'])}\n\n"
        )

    text += "Чтобы отметить заказ выполненным, используй команду: /done НОМЕР\nНапример: /done 1"
    bot.send_message(message.chat.id, text, parse_mode="HTML")


@bot.message_handler(commands=["done"])
def handle_done(message: types.Message):
    parts = message.text.split()

    if len(parts) != 2 or not parts[1].isdigit():
        bot.send_message(message.chat.id, "Использование: /done 2 (где 2 — номер заказа в очереди)")
        return

    idx = int(parts[1]) - 1

    if idx < 0 or idx >= len(orders_queue):
        bot.send_message(message.chat.id, "Неверный номер заказа.")
        return

    removed = orders_queue.pop(idx)
    bot.send_message(
        message.chat.id,
        f"✅ Заказ {removed['fio']} ({removed['phone']}) помечен как выполненный и удалён из очереди."
    )


@bot.message_handler(commands=["clearqueue"])
def handle_clear_queue(message: types.Message):
    orders_queue.clear()
    bot.send_message(message.chat.id, "Очередь заказов полностью очищена.")


@bot.message_handler(commands=["help"])
def handle_help(message: types.Message):
    bot.send_message(
        message.chat.id,
        "Доступные команды:\n"
        "/start — начать работу\n"
        "/help — помощь\n"
        "/queue — показать очередь заказов (для оператора)\n"
        "/done N — отметить заказ №N выполненным\n"
        "/clearqueue — очистить очередь заказов\n\n"
        "Основные действия доступны через кнопки: «Каталог товаров», «Корзина».",
        reply_markup=main_menu_keyboard(),
    )


# Фолбек на непонятные сообщения
@bot.message_handler(func=lambda m: True)
def handle_fallback(message: types.Message):
    bot.send_message(
        message.chat.id,
        "Я не понял сообщение.\n"
        "Используйте кнопки «Каталог товаров» или «Корзина», либо команду /help.",
        reply_markup=main_menu_keyboard(),
    )


if __name__ == "__main__":
    print("Бот запущен...")
    bot.infinity_polling()
